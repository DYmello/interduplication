from __future__ import annotations

import argparse
import logging
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

from src.embedding_model import BGEEmbeddingModel
from src.excel_schema import INPUT_HEADERS, InputRecord
from src.name_normalizer import clean_display_name
from src.similarity_engine import (
    SimilarityEngine,
    extract_objects,
    sort_results,
)
from src.xlsx_writer import write_similarity_workbook


LOGGER = logging.getLogger("detectduplication")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="使用 BGE-M3 dense embedding 检测软件和模块名称的文字语义相似性"
    )
    parser.add_argument("--input_files", type=Path, nargs="+", required=True)
    parser.add_argument(
        "--output_file", type=Path, default=Path("duplication_results.xlsx")
    )
    parser.add_argument("--comparison_mode", choices=["within", "cross", "all"], default="within")
    parser.add_argument(
        "--anchor_group",
        help="定向跨组比较的基准组编号；只适用于 cross 模式",
    )
    parser.add_argument(
        "--target_groups",
        type=parse_target_groups,
        help="定向跨组比较的目标组编号，多个组以逗号分隔；只适用于 cross 模式",
    )
    parser.add_argument("--levels", default="software,module")
    parser.add_argument("--similarity_threshold", type=float, default=0.85)
    parser.add_argument("--top_k", type=int, default=10)
    parser.add_argument("--include_unmatched", action="store_true")
    parser.add_argument("--remove_module_suffix", action="store_true")
    parser.add_argument("--model_name_or_path", default="BAAI/bge-m3")
    parser.add_argument("--batch_size", type=int, default=64)
    parser.add_argument("--max_length", type=int, default=64)
    parser.add_argument("--device", choices=["auto", "cuda", "cpu"], default="auto")
    parser.add_argument(
        "--use_fp16",
        action=argparse.BooleanOptionalAction,
        default=None,
        help="GPU 默认启用；可用 --no-use_fp16 显式关闭",
    )
    parser.add_argument("--block_size", type=int, default=2048)
    parser.add_argument("--overwrite", action="store_true")
    return parser


def parse_levels(value: str) -> list[str]:
    levels = [item.strip().lower() for item in value.split(",") if item.strip()]
    if not levels or any(item not in {"software", "module"} for item in levels):
        raise ValueError("levels 只能由 software,module 组成")
    result: list[str] = []
    for item in levels:
        if item not in result:
            result.append(item)
    return result


def parse_target_groups(value: str | None) -> list[str] | None:
    """Parse a comma-separated group list, preserving first-seen order."""
    if value is None:
        return None
    result: list[str] = []
    for item in value.split(","):
        group_id = item.strip()
        if group_id and group_id not in result:
            result.append(group_id)
    return result


def validate_directional_groups(
    comparison_mode: str,
    anchor_group: str | None,
    target_groups: Sequence[str] | None,
    available_groups: Sequence[str],
) -> tuple[str | None, list[str] | None]:
    """Validate and normalize the optional directed cross-group scope."""
    directional = anchor_group is not None or target_groups is not None
    if not directional:
        return None, None
    if comparison_mode != "cross":
        raise ValueError("anchor_group 和 target_groups 只适用于 cross 模式")
    if (anchor_group is None) != (target_groups is None):
        raise ValueError("anchor_group 和 target_groups 必须同时提供")

    normalized_anchor = anchor_group.strip() if anchor_group is not None else ""
    normalized_targets = list(target_groups or [])
    if not normalized_anchor:
        raise ValueError("anchor_group 不能为空")
    if not normalized_targets:
        raise ValueError("target_groups 解析后不能为空")

    available = set(available_groups)
    if normalized_anchor not in available:
        raise ValueError(f"anchor_group 不存在于输入数据中：{normalized_anchor}")
    unknown = [item for item in normalized_targets if item not in available]
    if unknown:
        raise ValueError(f"target_groups 中存在输入数据未包含的组编号：{','.join(unknown)}")
    if normalized_anchor in normalized_targets:
        raise ValueError("target_groups 不能包含 anchor_group")
    return normalized_anchor, normalized_targets


def read_standardized_files(paths: Iterable[Path]) -> list[InputRecord]:
    records: list[InputRecord] = []
    for path_value in paths:
        path = path_value.resolve()
        if not path.is_file():
            raise FileNotFoundError(f"输入文件不存在：{path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"输入文件必须是 .xlsx：{path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Input" not in workbook.sheetnames:
            raise ValueError(f"{path} 缺少 Input 工作表")
        worksheet = workbook["Input"]
        headers = [clean_display_name(worksheet.cell(1, col).value) for col in range(1, 10)]
        if headers != INPUT_HEADERS:
            raise ValueError(
                f"{path} 的 Input 表头或顺序错误：{headers}；期望：{INPUT_HEADERS}"
            )
        extra_headers = [
            clean_display_name(worksheet.cell(1, col).value)
            for col in range(10, worksheet.max_column + 1)
        ]
        if any(extra_headers):
            raise ValueError(f"{path} 的 Input 工作表存在未定义的额外列")
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, max_col=9, values_only=True),
            start=2,
        ):
            if all(value is None or clean_display_name(value) == "" for value in values):
                continue
            record = InputRecord.from_row(list(values))
            _validate_record(record, path, row_number)
            records.append(record)
        LOGGER.info("读取标准输入：%s，累计 %d 行", path, len(records))
    if not records:
        raise ValueError("标准输入中没有可用记录")
    _validate_software_entity_rows(records)
    return records


def _validate_record(record: InputRecord, path: Path, row_number: int) -> None:
    prefix = f"{path} 的 Input!第 {row_number} 行"
    required = {
        "组编号": record.group_id,
        "平台编号": record.platform_id,
        "平台名称": record.platform_name,
        "系统编号": record.system_id,
        "系统名称": record.system_name,
        "软件编号": record.software_id,
        "软件名称": record.software_name,
    }
    empty = [name for name, value in required.items() if not value]
    if empty:
        raise ValueError(f"{prefix} 缺少字段：{empty}")
    if bool(record.module_id) != bool(record.module_name):
        raise ValueError(f"{prefix} 的模块编号和模块名称必须同时为空或同时非空")


def _validate_software_entity_rows(records: list[InputRecord]) -> None:
    software_keys = {(item.group_id, item.software_id) for item in records}
    entity_keys = {
        (item.group_id, item.software_id) for item in records if not item.module_id
    }
    missing = sorted(software_keys - entity_keys)
    if missing:
        raise ValueError(f"以下软件缺少模块字段为空的软件实体行：{missing[:20]}")


def build_snapshot(objects_by_level: dict[str, list]) -> list[InputRecord]:
    rows: list[InputRecord] = []
    for item in objects_by_level["software"]:
        rows.append(
            InputRecord(
                item.group_id,
                item.platform_id,
                item.platform_name,
                item.system_id,
                item.system_name,
                item.software_id,
                item.software_name,
            )
        )
    for item in objects_by_level["module"]:
        rows.append(
            InputRecord(
                item.group_id,
                item.platform_id,
                item.platform_name,
                item.system_id,
                item.system_name,
                item.software_id,
                item.software_name,
                item.module_id,
                item.module_name,
            )
        )
    return sorted(
        rows,
        key=lambda item: (
            item.group_id,
            item.platform_id,
            item.system_id,
            item.software_id,
            0 if not item.module_id else 1,
            item.module_id,
        ),
    )


def run(args: argparse.Namespace) -> dict[str, int]:
    output_path = args.output_file.resolve()
    if output_path.exists() and not args.overwrite:
        raise FileExistsError(
            f"输出文件已存在：{output_path}；如需覆盖请添加 --overwrite"
        )
    levels = parse_levels(args.levels)
    if not 0 <= args.similarity_threshold <= 1:
        raise ValueError("similarity_threshold 必须位于 [0, 1]")
    if args.top_k <= 0 or args.batch_size <= 0 or args.max_length <= 0 or args.block_size <= 0:
        raise ValueError("top_k、batch_size、max_length 和 block_size 必须为正整数")

    records = read_standardized_files(args.input_files)
    objects_by_level = extract_objects(records)
    groups = []
    for record in records:
        if record.group_id not in groups:
            groups.append(record.group_id)
    anchor_group, target_groups = validate_directional_groups(
        args.comparison_mode,
        args.anchor_group,
        args.target_groups,
        groups,
    )
    if args.comparison_mode == "cross" and len(groups) < 2:
        raise ValueError("cross 模式至少需要两个不同组编号")

    embedder = BGEEmbeddingModel(
        args.model_name_or_path,
        device=args.device,
        use_fp16=args.use_fp16,
        logger=LOGGER,
    )
    engine = SimilarityEngine(embedder, LOGGER)
    results = []
    for level in levels:
        objects = objects_by_level[level]
        LOGGER.info("开始 %s 层比较：%d 个对象", level, len(objects))
        results.extend(
            engine.compare(
                objects,
                comparison_mode=args.comparison_mode,
                similarity_threshold=args.similarity_threshold,
                top_k=args.top_k,
                include_unmatched=args.include_unmatched,
                remove_module_suffix=args.remove_module_suffix,
                batch_size=args.batch_size,
                max_length=args.max_length,
                block_size=args.block_size,
                anchor_group=anchor_group,
                target_groups=target_groups,
            )
        )
    results = sort_results(results)
    snapshot = build_snapshot(objects_by_level)
    config = {
        "model_name_or_path": args.model_name_or_path,
        "comparison_mode": args.comparison_mode,
        "anchor_group": anchor_group or "",
        "target_groups": ",".join(target_groups or []),
        "levels": ",".join(levels),
        "same_level_only": True,
        "similarity_threshold": args.similarity_threshold,
        "top_k": args.top_k,
        "include_unmatched": args.include_unmatched,
        "remove_module_suffix": args.remove_module_suffix,
        "batch_size": args.batch_size,
        "max_length": args.max_length,
        "device": embedder.device,
        "use_fp16": embedder.use_fp16,
        "similarity_metric": "cosine",
    }
    write_similarity_workbook(output_path, results, snapshot, config)
    counts = {
        "software": len(objects_by_level["software"]),
        "module": len(objects_by_level["module"]),
        "results": len(results),
    }
    LOGGER.info(
        "检测完成：软件=%d，模块=%d，候选对=%d，输出=%s",
        counts["software"],
        counts["module"],
        counts["results"],
        output_path,
    )
    return counts


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    args = build_parser().parse_args()
    try:
        run(args)
    except (OSError, RuntimeError, ValueError) as exc:
        LOGGER.error("检测失败：%s", exc)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
