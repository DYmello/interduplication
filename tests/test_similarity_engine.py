from __future__ import annotations

import numpy as np
from openpyxl import load_workbook

from src.excel_schema import INPUT_HEADERS, SIMILARITY_HEADERS, InputRecord
from src.similarity_engine import NameObject, SimilarityEngine
from src.xlsx_writer import write_similarity_workbook


class MockEmbedder:
    def __init__(self, vectors: dict[str, list[float]]):
        self.vectors = vectors

    def encode(self, texts, *, batch_size, max_length):
        return np.asarray([self.vectors[text] for text in texts], dtype=np.float32)


def software(group: str, object_id: str, name: str) -> NameObject:
    return NameObject(
        "software", group, "P001", "平台", "S001", "系统", object_id, name
    )


def module(group: str, object_id: str, name: str) -> NameObject:
    return NameObject(
        "module",
        group,
        "P001",
        "平台",
        "S001",
        "系统",
        "SW001",
        "软件",
        object_id,
        name,
    )


def compare(
    engine,
    objects,
    mode="within",
    threshold=0.85,
    *,
    top_k=10,
    include_unmatched=False,
    anchor_group=None,
    target_groups=None,
):
    return engine.compare(
        objects,
        comparison_mode=mode,
        similarity_threshold=threshold,
        top_k=top_k,
        include_unmatched=include_unmatched,
        remove_module_suffix=False,
        batch_size=8,
        max_length=64,
        block_size=2,
        anchor_group=anchor_group,
        target_groups=target_groups,
    )


def test_self_cosine_similarity_is_one():
    vector = np.asarray([[3.0, 4.0]], dtype=np.float32)
    vector /= np.linalg.norm(vector, axis=1, keepdims=True)
    assert float(vector[0] @ vector[0]) == 1.0


def test_within_has_no_self_or_symmetric_duplicates():
    embedder = MockEmbedder({"同名软件": [1, 0], "其他软件": [0.9, 0.1]})
    objects = [
        software("A", "SW001", "同名软件"),
        software("A", "SW002", "同名软件"),
        software("A", "SW003", "其他软件"),
    ]
    results = compare(SimilarityEngine(embedder), objects)
    identities = [(r.object_a.identity, r.object_b.identity) for r in results]
    assert all(left != right for left, right in identities)
    unordered = {frozenset((left, right)) for left, right in identities}
    assert len(unordered) == len(identities)


def test_cross_only_compares_different_groups():
    embedder = MockEmbedder({"同名软件": [1, 0]})
    objects = [
        software("A", "SW001", "同名软件"),
        software("A", "SW002", "同名软件"),
        software("B", "SW001", "同名软件"),
    ]
    results = compare(SimilarityEngine(embedder), objects, mode="cross")
    assert results
    assert all(r.object_a.group_id != r.object_b.group_id for r in results)
    assert all(r.object_a.group_id == "A" for r in results)


def test_directed_cross_only_generates_anchor_to_selected_targets():
    embedder = MockEmbedder({"same": [1, 0]})
    objects = [
        software("B", "B01", "same"),
        software("A", "A01", "same"),
        software("C", "C01", "same"),
        software("D", "D01", "same"),
    ]
    results = compare(
        SimilarityEngine(embedder),
        objects,
        mode="cross",
        anchor_group="A",
        target_groups=["B", "C", "D"],
    )
    pairs = {(item.object_a.group_id, item.object_b.group_id) for item in results}
    assert pairs == {("A", "B"), ("A", "C"), ("A", "D")}
    assert all(item.object_a.group_id == "A" for item in results)
    assert all(item.object_b.group_id in {"B", "C", "D"} for item in results)


def test_directed_cross_ranks_all_target_groups_together_and_applies_total_top_k():
    embedder = MockEmbedder(
        {
            "anchor": [1.0, 0.0],
            "best": [1.0, 0.0],
            "second": [0.9, 0.43589],
            "third": [0.8, 0.6],
        }
    )
    objects = [
        software("A", "A01", "anchor"),
        software("B", "B01", "third"),
        software("C", "C01", "best"),
        software("D", "D01", "second"),
    ]
    results = compare(
        SimilarityEngine(embedder),
        objects,
        mode="cross",
        threshold=0.0,
        top_k=2,
        anchor_group="A",
        target_groups=["B", "C", "D"],
    )
    assert [(item.object_b.group_id, item.rank) for item in results] == [
        ("C", 1),
        ("D", 2),
    ]


def test_directed_cross_tie_breaks_by_target_group_then_object_id():
    embedder = MockEmbedder({"same": [1, 0]})
    objects = [
        software("A", "A01", "same"),
        software("C", "C02", "same"),
        software("B", "B02", "same"),
        software("B", "B01", "same"),
    ]
    results = compare(
        SimilarityEngine(embedder),
        objects,
        mode="cross",
        anchor_group="A",
        target_groups=["C", "B"],
    )
    assert [(item.object_b.group_id, item.object_b.object_id) for item in results] == [
        ("B", "B01"),
        ("B", "B02"),
        ("C", "C02"),
    ]


def test_directed_cross_include_unmatched_only_outputs_anchor_objects():
    embedder = MockEmbedder({"anchor": [1, 0], "target": [0, 1]})
    objects = [
        software("B", "B01", "target"),
        software("A", "A01", "anchor"),
        software("C", "C01", "target"),
    ]
    results = compare(
        SimilarityEngine(embedder),
        objects,
        mode="cross",
        threshold=0.9,
        include_unmatched=True,
        anchor_group="A",
        target_groups=["B", "C"],
    )
    assert len(results) == 1
    assert results[0].object_a.group_id == "A"
    assert results[0].object_b is None


def test_cross_without_directional_parameters_keeps_all_group_combinations():
    embedder = MockEmbedder({"same": [1, 0]})
    objects = [
        software("A", "A01", "same"),
        software("B", "B01", "same"),
        software("C", "C01", "same"),
    ]
    results = compare(SimilarityEngine(embedder), objects, mode="cross")
    pairs = {(item.object_a.group_id, item.object_b.group_id) for item in results}
    assert pairs == {("A", "B"), ("A", "C"), ("B", "C")}


def test_identical_normalized_names_are_labeled_exact():
    embedder = MockEmbedder({"同名软件": [1, 0]})
    results = compare(
        SimilarityEngine(embedder),
        [software("A", "SW001", "同名软件"), software("A", "SW002", "同名软件")],
    )
    assert len(results) == 1
    assert results[0].similarity_level == "名称完全相同"


def test_cosine_rounding_is_clamped_to_one():
    embedder = MockEmbedder({"同名软件": [1.0] * 7})
    results = compare(
        SimilarityEngine(embedder),
        [software("A", "SW001", "同名软件"), software("A", "SW002", "同名软件")],
    )
    assert results[0].score == 1.0


def test_below_threshold_is_not_output():
    embedder = MockEmbedder({"软件甲": [1, 0], "软件乙": [0, 1]})
    results = compare(
        SimilarityEngine(embedder),
        [software("A", "SW001", "软件甲"), software("A", "SW002", "软件乙")],
    )
    assert results == []


def test_software_and_module_are_compared_in_separate_calls():
    embedder = MockEmbedder({"同名": [1, 0]})
    engine = SimilarityEngine(embedder)
    software_results = compare(
        engine,
        [software("A", "SW001", "同名"), software("A", "SW002", "同名")],
    )
    module_results = compare(
        engine,
        [module("A", "M0001", "同名"), module("A", "M0002", "同名")],
    )
    assert {item.level for item in software_results} == {"software"}
    assert {item.level for item in module_results} == {"module"}


def test_similarity_workbook_column_order(tmp_path):
    output = tmp_path / "result.xlsx"
    write_similarity_workbook(
        output,
        [],
        [InputRecord("A", "P001", "平台", "S001", "系统", "SW001", "软件")],
        {"similarity_metric": "cosine", "same_level_only": True},
    )
    workbook = load_workbook(output, read_only=True, data_only=True)
    result_headers = [cell.value for cell in workbook["SimilarityResults"][1]]
    snapshot_headers = [cell.value for cell in workbook["InputSnapshot"][1]]
    assert result_headers == SIMILARITY_HEADERS
    assert snapshot_headers == INPUT_HEADERS


def test_similarity_workbook_records_actual_directional_scope(tmp_path):
    output = tmp_path / "result.xlsx"
    write_similarity_workbook(
        output,
        [],
        [],
        {"anchor_group": "A", "target_groups": "B,C,D"},
    )
    workbook = load_workbook(output, read_only=True, data_only=True)
    config = {
        row[0]: (row[1], row[2])
        for row in workbook["Config"].iter_rows(min_row=2, values_only=True)
    }
    assert config["anchor_group"] == ("A", "定向跨组比较的基准组")
    assert config["target_groups"] == ("B,C,D", "定向跨组比较的目标组")
