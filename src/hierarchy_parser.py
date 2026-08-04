from __future__ import annotations

import logging
from dataclasses import dataclass
from typing import Iterable

from openpyxl.worksheet.worksheet import Worksheet

from .excel_schema import ExcludedCategory, InputRecord
from .name_normalizer import classify_serial, clean_display_name, normalize_serial


class HierarchyError(ValueError):
    """Raised when the source hierarchy is invalid in strict mode."""


@dataclass(frozen=True)
class ParseResult:
    records: list[InputRecord]
    excluded_categories: list[ExcludedCategory]
    counts: dict[str, int]
    warnings: list[str]


def find_header_row(worksheet: Worksheet, scan_limit: int = 30) -> int:
    for row_number in range(1, min(worksheet.max_row, scan_limit) + 1):
        serial_header = clean_display_name(worksheet.cell(row_number, 1).value)
        name_header = clean_display_name(worksheet.cell(row_number, 2).value)
        if serial_header == "序号" and name_header == "应用系统名称":
            return row_number
    raise HierarchyError("未找到表头：第一、二列必须包含“序号”和“应用系统名称”")


def parse_hierarchy(
    worksheet: Worksheet,
    *,
    group_id: str = "A",
    strict: bool = False,
    logger: logging.Logger | None = None,
) -> ParseResult:
    log = logger or logging.getLogger(__name__)
    normalized_group = clean_display_name(group_id)
    if not normalized_group:
        raise HierarchyError("组编号不能为空")
    if worksheet.max_column < 2:
        raise HierarchyError("原始工作表至少需要两列")

    header_row = find_header_row(worksheet)
    counters = {"platform": 0, "system": 0, "software": 0, "module": 0}
    records: list[InputRecord] = []
    excluded: list[ExcludedCategory] = []
    warnings: list[str] = []

    current_platform = ("", "")
    current_system = ("", "")
    current_software = ("", "")

    def structural_issue(row_number: int, message: str) -> None:
        full_message = f"第 {row_number} 行：{message}"
        if strict:
            raise HierarchyError(full_message)
        warnings.append(full_message)
        log.warning(full_message)

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        raw_serial = worksheet.cell(row_number, 1).value
        raw_name = worksheet.cell(row_number, 2).value
        serial = normalize_serial(raw_serial)
        name = clean_display_name(raw_name)

        if not serial and not name:
            continue
        if serial == "合计" or name == "合计":
            continue
        if not name:
            structural_issue(row_number, f"序号 {serial!r} 对应的名称为空，已跳过")
            continue

        level = classify_serial(serial, name)
        if level == "platform":
            counters["platform"] += 1
            current_platform = (f"P{counters['platform']:03d}", name)
            current_system = ("", "")
            current_software = ("", "")
            continue

        if level == "system":
            if not current_platform[0]:
                structural_issue(row_number, "系统行缺少当前平台")
            counters["system"] += 1
            current_system = (f"S{counters['system']:03d}", name)
            current_software = ("", "")
            continue

        if level == "software":
            if not current_platform[0] or not current_system[0]:
                structural_issue(row_number, "软件行缺少当前平台或系统")
            counters["software"] += 1
            current_software = (f"SW{counters['software']:03d}", name)
            records.append(
                InputRecord(
                    normalized_group,
                    current_platform[0],
                    current_platform[1],
                    current_system[0],
                    current_system[1],
                    current_software[0],
                    current_software[1],
                )
            )
            continue

        if level == "module":
            if not current_software[0]:
                structural_issue(row_number, "模块行缺少当前软件，已跳过")
                continue
            counters["module"] += 1
            records.append(
                InputRecord(
                    normalized_group,
                    current_platform[0],
                    current_platform[1],
                    current_system[0],
                    current_system[1],
                    current_software[0],
                    current_software[1],
                    f"M{counters['module']:04d}",
                    name,
                )
            )
            continue

        if level == "category":
            excluded.append(
                ExcludedCategory(
                    current_platform[0],
                    current_platform[1],
                    current_system[0],
                    current_system[1],
                    name,
                    row_number,
                )
            )
            continue

        structural_issue(row_number, f"无法识别序号格式 {serial!r}，已跳过")

    _validate_result(records, counters)
    return ParseResult(records, excluded, counters, warnings)


def _validate_result(records: Iterable[InputRecord], counts: dict[str, int]) -> None:
    software_ids: set[str] = set()
    module_ids: set[str] = set()
    software_entity_rows = 0
    for record in records:
        if not record.software_id or not record.software_name:
            raise HierarchyError("标准记录中存在空软件编号或软件名称")
        if record.module_id:
            if not record.module_name:
                raise HierarchyError(f"模块 {record.module_id} 的名称为空")
            if record.module_id in module_ids:
                raise HierarchyError(f"模块编号重复：{record.module_id}")
            module_ids.add(record.module_id)
        else:
            if record.module_name:
                raise HierarchyError("软件实体行的模块名称必须为空")
            if record.software_id in software_ids:
                raise HierarchyError(f"软件编号重复：{record.software_id}")
            software_ids.add(record.software_id)
            software_entity_rows += 1
    if software_entity_rows != counts["software"]:
        raise HierarchyError("软件计数与软件实体行数不一致")
    if len(module_ids) != counts["module"]:
        raise HierarchyError("模块计数与模块记录数不一致")
