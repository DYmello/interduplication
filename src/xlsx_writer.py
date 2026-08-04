from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .excel_schema import (
    EXCLUDED_HEADERS,
    FORMTRANS_CONFIG,
    INPUT_HEADERS,
    SIMILARITY_HEADERS,
    ExcludedCategory,
    InputRecord,
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))
TEXT_ALIGNMENT = Alignment(vertical="top", wrap_text=False)


def _prepare_sheet(worksheet: Worksheet, headers: Sequence[str]) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = SUBTLE_BORDER
    worksheet.row_dimensions[1].height = 24


def _finish_sheet(
    worksheet: Worksheet,
    widths: Sequence[float],
    *,
    filter_enabled: bool = True,
) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    if filter_enabled and worksheet.max_row >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = TEXT_ALIGNMENT


def _save_atomic(workbook: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(f".{output_path.stem}.tmp.xlsx")
    try:
        workbook.save(temporary)
        temporary.replace(output_path)
    finally:
        if temporary.exists():
            temporary.unlink()


def write_normalized_workbook(
    output_path: Path,
    records: Iterable[InputRecord],
    excluded_categories: Iterable[ExcludedCategory],
) -> None:
    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = "Input"
    _prepare_sheet(input_sheet, INPUT_HEADERS)
    for record in records:
        input_sheet.append(record.as_row())
    _finish_sheet(input_sheet, [12, 12, 24, 12, 26, 13, 34, 13, 38])

    config_sheet = workbook.create_sheet("Config")
    _prepare_sheet(config_sheet, ["参数", "值", "说明"])
    for key, value, description in FORMTRANS_CONFIG:
        config_sheet.append([key, value, description])
    _finish_sheet(config_sheet, [28, 22, 54], filter_enabled=False)

    excluded_sheet = workbook.create_sheet("Excluded_Categories")
    _prepare_sheet(excluded_sheet, EXCLUDED_HEADERS)
    for item in excluded_categories:
        excluded_sheet.append(item.as_row())
    _finish_sheet(excluded_sheet, [12, 24, 12, 26, 38, 12])

    _save_atomic(workbook, output_path)


def _result_row(sequence: int, result: Any) -> list[Any]:
    a = result.object_a
    b = result.object_b
    if b is None:
        b_values = [""] * 9
    else:
        b_values = [
            b.group_id,
            b.platform_id,
            b.platform_name,
            b.system_id,
            b.system_name,
            b.software_id,
            b.software_name,
            b.module_id if b.level == "module" else "",
            b.module_name if b.level == "module" else "",
        ]
    return [
        sequence,
        result.level,
        result.rank,
        result.score if result.score is not None else "",
        result.similarity_level,
        a.group_id,
        a.platform_id,
        a.platform_name,
        a.system_id,
        a.system_name,
        a.software_id,
        a.software_name,
        a.module_id if a.level == "module" else "",
        a.module_name if a.level == "module" else "",
        *b_values,
    ]


def write_similarity_workbook(
    output_path: Path,
    results: Sequence[Any],
    input_snapshot: Sequence[InputRecord],
    config: dict[str, Any],
) -> None:
    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "SimilarityResults"
    _prepare_sheet(result_sheet, SIMILARITY_HEADERS)
    for sequence, result in enumerate(results, start=1):
        result_sheet.append(_result_row(sequence, result))
    for cell in result_sheet["D"][1:]:
        cell.number_format = "0.000000"
    _finish_sheet(
        result_sheet,
        [8, 12, 12, 20, 16, 10, 12, 24, 12, 26, 13, 34, 13, 38,
         10, 12, 24, 12, 26, 13, 34, 13, 38],
    )

    snapshot_sheet = workbook.create_sheet("InputSnapshot")
    _prepare_sheet(snapshot_sheet, INPUT_HEADERS)
    for record in input_snapshot:
        snapshot_sheet.append(record.as_row())
    _finish_sheet(snapshot_sheet, [12, 12, 24, 12, 26, 13, 34, 13, 38])

    config_sheet = workbook.create_sheet("Config")
    _prepare_sheet(config_sheet, ["参数", "值", "说明"])
    descriptions = {
        "anchor_group": "定向跨组比较的基准组",
        "target_groups": "定向跨组比较的目标组",
        "similarity_threshold": "初始工程阈值，正式使用前应基于人工标注名称对校准",
        "similarity_metric": "L2 归一化稠密向量的余弦相似度，不是概率或重复率",
        "same_level_only": "固定为 true，软件与模块不得跨层比较",
    }
    for key, value in config.items():
        if isinstance(value, bool):
            value = str(value).lower()
        config_sheet.append([key, value, descriptions.get(key, "本次实际运行参数")])
    _finish_sheet(config_sheet, [28, 38, 64], filter_enabled=False)

    _save_atomic(workbook, output_path)
