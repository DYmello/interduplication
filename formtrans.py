from __future__ import annotations

import argparse
import logging
from pathlib import Path

from openpyxl import load_workbook

from src.hierarchy_parser import HierarchyError, parse_hierarchy
from src.xlsx_writer import write_normalized_workbook


LOGGER = logging.getLogger("formtrans")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="将两列层级表格转换为名称重复性检测标准输入 XLSX"
    )
    parser.add_argument("--input_file", type=Path, required=True)
    parser.add_argument(
        "--output_file", type=Path, default=Path("normalized_input.xlsx")
    )
    parser.add_argument("--group_id", default="A")
    parser.add_argument("--sheet_name")
    parser.add_argument("--strict", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    return parser


def run(args: argparse.Namespace) -> dict[str, int]:
    input_path = args.input_file.resolve()
    output_path = args.output_file.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("input_file 必须是 .xlsx 文件")
    if output_path.exists() and not args.overwrite:
        raise FileExistsError(
            f"输出文件已存在：{output_path}；如需覆盖请添加 --overwrite"
        )
    if input_path == output_path:
        raise ValueError("输入文件和输出文件不能是同一路径")

    workbook = load_workbook(input_path, data_only=True, read_only=False)
    if args.sheet_name:
        if args.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"工作表不存在：{args.sheet_name}；可用工作表：{workbook.sheetnames}"
            )
        worksheet = workbook[args.sheet_name]
    else:
        worksheet = workbook.worksheets[0]
    LOGGER.info("读取工作表：%s", worksheet.title)

    result = parse_hierarchy(
        worksheet,
        group_id=args.group_id,
        strict=args.strict,
        logger=LOGGER,
    )
    write_normalized_workbook(
        output_path,
        result.records,
        result.excluded_categories,
    )
    counts = {
        **result.counts,
        "excluded_categories": len(result.excluded_categories),
        "output_rows": len(result.records),
        "warnings": len(result.warnings),
    }
    LOGGER.info(
        "转换完成：平台=%d，系统=%d，软件=%d，模块=%d，分类标题=%d，输出=%s",
        counts["platform"],
        counts["system"],
        counts["software"],
        counts["module"],
        counts["excluded_categories"],
        output_path,
    )
    return counts


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    args = build_parser().parse_args()
    try:
        run(args)
    except (OSError, ValueError, HierarchyError) as exc:
        LOGGER.error("转换失败：%s", exc)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
